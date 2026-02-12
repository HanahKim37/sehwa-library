# pages/02_독서가.py

import re
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from difflib import SequenceMatcher

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# =========================================================
# 0) 임계값
# =========================================================
REQUIRED_TITLE_THRESHOLD = 0.80
REQUIRED_AUTHOR_THRESHOLD = 0.60
DUP_TITLE_THRESHOLD = 0.80
DUP_AUTHOR_THRESHOLD = 0.60

# =========================================================
# 1) 페이지 설정
# =========================================================
APP_TITLE = "📚 독서가"
st.set_page_config(page_title="독서가", page_icon="📚", layout="wide")
st.title(APP_TITLE)
st.caption("독서활동상황 엑셀을 업로드하면 학기별 충족 여부와 총 충족 여부를 산출합니다.")

st.info(
    "※ 독서활동상황 엑셀은 **한 반의 파일만** 업로드해 주세요.\n\n"
    "✅ 또한 **파일명에 학년-반을 반드시 포함**해 주세요. (예: `1-1`, `3-12`)\n"
    "- 예시 파일명: `독서활동_2-10.xlsx`, `2-10_독서활동상황.xlsx`"
)

# =========================================================
# 2) 내장 필독서 경로
# =========================================================
PROJECT_ROOT = Path(__file__).resolve().parents[1]
REQUIRED_2024_PATH = PROJECT_ROOT / "data" / "required_books" / "필독서_2024.xlsx"
REQUIRED_2025_PATH = PROJECT_ROOT / "data" / "required_books" / "필독서_2025.xlsx"

# =========================================================
# 3) 유틸
# =========================================================
def _normalize_text(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[^\w가-힣]", "", s)
    return s


def _title_variants_norm(title: str) -> List[str]:
    if title is None:
        return []
    t = str(title).strip()
    if not t:
        return []
    seps = ["(", "（", "[", "【", ":", "：", "/", "／", " - ", " – ", " — ", "-", "–", "—"]
    candidates = [t]
    for sep in seps:
        if sep in t:
            candidates.append(t.split(sep, 1)[0].strip())

    out: List[str] = []
    for c in candidates:
        n = _normalize_text(c)
        if n and n not in out:
            out.append(n)
    return out


def _author_variants_norm(author: str) -> List[str]:
    if author is None:
        return []
    a = str(author).strip()
    if not a:
        return []
    tokens = [t for t in re.split(r"\s+", a) if t.strip()]
    out: List[str] = []

    full = _normalize_text(a)
    if full:
        out.append(full)

    if tokens:
        last = tokens[-1]
        last_n = _normalize_text(last)
        if last_n and last_n not in out:
            out.append(last_n)

        initials = []
        for t in tokens[:-1]:
            t2 = re.sub(r"[^A-Za-z가-힣]", "", t)
            if t2:
                initials.append(t2[0].lower())
        if initials and last_n:
            combo = _normalize_text("".join(initials) + last)
            if combo and combo not in out:
                out.append(combo)

    return out


def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()


def _best_title_similarity(title_a: str, title_b: str) -> float:
    va = _title_variants_norm(title_a)
    vb = _title_variants_norm(title_b)
    best = 0.0
    for x in va:
        for y in vb:
            if len(x) < 3 or len(y) < 3:
                continue
            best = max(best, _similarity(x, y))
    return best


def _best_author_similarity(author_a: str, author_b: str) -> float:
    va = _author_variants_norm(author_a)
    vb = _author_variants_norm(author_b)
    best = 0.0
    for x in va:
        for y in vb:
            if len(x) < 3 or len(y) < 3:
                continue
            if (len(x) >= 4 and x in y) or (len(y) >= 4 and y in x):
                best = max(best, 1.0)
            else:
                best = max(best, _similarity(x, y))
    return best


def _pick_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    cols = list(df.columns)
    for cand in candidates:
        for col in cols:
            if col is None:
                continue
            if cand in str(col).replace(" ", ""):
                return col
    return None


def _extract_grade_class_from_filename_with_raw(filename: str) -> Tuple[Optional[int], Optional[int], Optional[str]]:
    if not filename:
        return None, None, None
    m = re.search(r"(?<!\d)([1-3])\s*[-_ ]\s*(\d{1,2})(?!\d)", filename)
    if not m:
        return None, None, None
    g = int(m.group(1))
    c_str = m.group(2)
    c = int(c_str)
    if g not in [1, 2, 3]:
        return None, None, None
    if not (1 <= c <= 12):
        return None, None, None
    raw = f"{g}-{c_str}"
    return g, c, raw


def _parse_grade_class_from_sheet_top(df_raw: pd.DataFrame) -> Tuple[Optional[int], Optional[int]]:
    max_r = min(20, df_raw.shape[0])
    max_c = min(10, df_raw.shape[1])
    pattern = re.compile(r"(\d)\s*학\s*년\D{0,10}?(\d{1,2})\s*반")
    for r in range(max_r):
        for c in range(max_c):
            v = df_raw.iat[r, c]
            if pd.isna(v):
                continue
            m = pattern.search(str(v))
            if not m:
                continue
            grade = int(m.group(1))
            cls = int(m.group(2))
            if grade in [1, 2, 3] and cls >= 1:
                return grade, cls
    return None, None


def _infer_academic_year_from_print_date(df_raw_top: pd.DataFrame) -> Optional[int]:
    date_pat = re.compile(r"(20\d{2})[.\-/]\s*(\d{1,2})[.\-/]\s*(\d{1,2})")
    max_r = min(30, df_raw_top.shape[0])
    max_c = min(20, df_raw_top.shape[1])
    for r in range(max_r):
        for c in range(max_c):
            v = df_raw_top.iat[r, c]
            if pd.isna(v):
                continue
            m = date_pat.search(str(v))
            if m:
                y = int(m.group(1))
                mo = int(m.group(2))
                return (y - 1) if mo in [1, 2] else y
    return None


def _find_header_row_for_reading(df_raw: pd.DataFrame) -> Optional[int]:
    for r in range(min(80, df_raw.shape[0])):
        row_vals = [str(x).strip() for x in df_raw.iloc[r].tolist() if not pd.isna(x)]
        joined = " ".join(row_vals).replace(" ", "")
        if "번호" in joined and ("독서" in joined or "독서활동" in joined or "독서활동상황" in joined):
            return r
    return None


def _load_reading_table(xls: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    df_raw = xls.parse(sheet_name=sheet_name, header=None, dtype=str)
    header_row = _find_header_row_for_reading(df_raw)
    if header_row is None:
        return xls.parse(sheet_name=sheet_name, dtype=str)

    headers = df_raw.iloc[header_row].tolist()
    df = df_raw.iloc[header_row + 1 :].copy()
    df.columns = headers
    df = df.dropna(how="all")
    return df


BOOK_PAIR_RE = re.compile(r"(?P<title>[^()]+?)\s*\(\s*(?P<author>[^()]+?)\s*\)")


def _split_books(cell_text: str) -> List[Tuple[str, str]]:
    if cell_text is None or pd.isna(cell_text):
        return []
    text = str(cell_text).strip()
    if not text:
        return []
    books: List[Tuple[str, str]] = []
    for m in BOOK_PAIR_RE.finditer(text):
        title = m.group("title").strip().rstrip(",").strip()
        author = m.group("author").strip().rstrip(",").strip()
        if title and author:
            books.append((title, author))
    return books


def _semester_sort_key(year: str, sem: str) -> Tuple[int, int]:
    try:
        y = int(re.findall(r"\d+", str(year))[0])
    except Exception:
        y = 9999
    nums = re.findall(r"\d+", str(sem))
    s = int(nums[0]) if nums else 9
    return (y, s)


def _find_req_header_row(df_raw: pd.DataFrame) -> Optional[int]:
    max_r = min(200, df_raw.shape[0])
    for r in range(max_r):
        row = df_raw.iloc[r].astype(str).fillna("").tolist()
        joined = " ".join([x.strip() for x in row if x and x.strip() != "nan"]).replace(" ", "")
        if ("도서명" in joined) and (("저자명" in joined) or ("저자" in joined)):
            return r
    return None


@st.cache_data(show_spinner=False)
def _load_required_books_from_repo(xlsx_path: str) -> Dict[str, Tuple[str, str]]:
    p = Path(xlsx_path)
    if not p.exists():
        raise FileNotFoundError(f"필독서 파일을 찾을 수 없습니다: {p.resolve()}")

    xls = pd.ExcelFile(p, engine="openpyxl")
    sheet = xls.sheet_names[0]

    raw = xls.parse(sheet_name=sheet, header=None, dtype=str).dropna(how="all")
    header_row = _find_req_header_row(raw)
    if header_row is None:
        raise ValueError(f"필독서 파일에서 '도서명/저자(저자명)' 헤더 행을 찾지 못했습니다: {p.resolve()}")

    headers = raw.iloc[header_row].tolist()
    df = raw.iloc[header_row + 1 :].copy()
    df.columns = headers

    title_col = _pick_col(df, ["도서명", "도서", "제목"])
    author_col = _pick_col(df, ["저자명", "저자", "지은이", "작가"])
    if not title_col or not author_col:
        raise ValueError(f"필독서 파일 컬럼 인식 실패: {p.resolve()}")

    out: Dict[str, Tuple[str, str]] = {}
    for _, row in df.iterrows():
        t = row.get(title_col, None)
        a = row.get(author_col, None)
        if t is None or a is None:
            continue
        if pd.isna(t) or pd.isna(a):
            continue

        t = str(t).strip()
        a = str(a).strip()
        if not t or not a:
            continue

        t_ns = t.replace(" ", "")
        a_ns = a.replace(" ", "")
        if t_ns in ["도서명", "도서", "제목"] or a_ns in ["저자", "저자명", "지은이"]:
            continue
        if t.lower() == "nan" or a.lower() == "nan":
            continue

        key = _normalize_text(t) + "|" + _normalize_text(a)
        out[key] = (t, a)

    if not out:
        raise ValueError(f"필독서 데이터가 0건으로 인식되었습니다: {p.resolve()}")

    return out


def _safe_year_int(x) -> Optional[int]:
    try:
        return int(re.findall(r"\d+", str(x))[0])
    except Exception:
        return None


def _choose_required_map(year_value, req_2024_map, req_2025_map):
    y = _safe_year_int(year_value)
    if y == 2024:
        return req_2024_map
    if y is None or y >= 2025:
        return req_2025_map
    return req_2024_map


def _build_required_title_map(req_map: Dict[str, Tuple[str, str]]) -> Dict[str, Tuple[str, str]]:
    title_map: Dict[str, Tuple[str, str]] = {}
    for _, (t, a) in req_map.items():
        for nt in _title_variants_norm(t):
            if nt not in title_map:
                title_map[nt] = (t, a)
    return title_map


def _fuzzy_match_required_title_author(
    raw_title: str,
    raw_author: str,
    req_title_map: Dict[str, Tuple[str, str]],
    title_threshold: float = REQUIRED_TITLE_THRESHOLD,
    author_threshold: float = REQUIRED_AUTHOR_THRESHOLD,
) -> Optional[Tuple[str, Tuple[str, str], float, float]]:
    if not raw_title:
        return None

    best_key = None
    best_title_score = 0.0
    best_std = None

    for k, v in req_title_map.items():
        if not k or len(k) < 3:
            continue
        title_scores = []
        for vt in _title_variants_norm(raw_title):
            if len(vt) < 3:
                continue
            title_scores.append(_similarity(vt, k))
        if not title_scores:
            continue
        score = max(title_scores)
        if score > best_title_score:
            best_title_score = score
            best_key = k
            best_std = v

    if best_key is None or best_std is None or best_title_score < title_threshold:
        return None

    std_t, std_a = best_std
    author_score = _best_author_similarity(raw_author, std_a)
    if author_score < author_threshold:
        return None

    return best_key, (std_t, std_a), best_title_score, author_score


def _find_duplicate_against_seen(
    title: str,
    author: str,
    seen: List[dict],
    title_threshold: float = DUP_TITLE_THRESHOLD,
    author_threshold: float = DUP_AUTHOR_THRESHOLD,
) -> Optional[Tuple[dict, float, float]]:
    best_item = None
    best_title = 0.0
    best_author = 0.0

    for it in seen:
        t2 = it.get("title", "")
        a2 = it.get("author", "")

        ts = _best_title_similarity(title, t2)
        if ts < title_threshold:
            continue

        ascore = _best_author_similarity(author, a2)
        if ascore < author_threshold:
            continue

        if (ts > best_title) or (ts == best_title and ascore > best_author):
            best_item = it
            best_title = ts
            best_author = ascore

    if best_item is None:
        return None
    return best_item, best_title, best_author


# openpyxl rich text 가능 여부
RICH_TEXT_AVAILABLE = False
try:
    from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont  # type: ignore
    RICH_TEXT_AVAILABLE = True
except Exception:
    RICH_TEXT_AVAILABLE = False


def _set_books_cell(ws, row_idx: int, col_idx: int, books: List[Tuple[str, str, bool]]):
    if not books:
        ws.cell(row=row_idx, column=col_idx).value = ""
        return

    cleaned = []
    for (t, a, is_req) in books:
        tt = str(t).strip().rstrip(",").strip()
        aa = str(a).strip().rstrip(",").strip()
        cleaned.append((tt, aa, is_req))

    if RICH_TEXT_AVAILABLE:
        rt = CellRichText()
        for i, (t, a, is_req) in enumerate(cleaned):
            text = f"{t}({a})"
            font = InlineFont(b=True) if is_req else InlineFont(b=False)
            rt.append(TextBlock(font, text))
            if i != len(cleaned) - 1:
                rt.append(TextBlock(InlineFont(b=False), ", "))
        ws.cell(row=row_idx, column=col_idx).value = rt
    else:
        items = []
        for (t, a, is_req) in cleaned:
            prefix = "★" if is_req else ""
            items.append(f"{prefix}{t}({a})")
        ws.cell(row=row_idx, column=col_idx).value = ", ".join(items)


def _sem_num(s: str) -> Optional[int]:
    nums = re.findall(r"\d+", str(s))
    if not nums:
        return None
    n = int(nums[0])
    return n if n in [1, 2] else None


def _row_semkey(row: dict, base_year: Optional[int], next_year: Optional[int]) -> Optional[str]:
    y = _safe_year_int(row.get("학년도"))
    s = _sem_num(row.get("학기"))
    if y is None or s is None or base_year is None:
        return None
    if y == base_year:
        return f"1{s}"
    if next_year is not None and y == next_year:
        return f"2{s}"
    return None


# =========================================================
# 4) 업로드/실행 UI
# =========================================================
reading_file = st.file_uploader("독서활동상황 엑셀 업로드 (.xlsx)", type=["xlsx"], key="reading_file")
run = st.button("✅ 충족 여부 판단 실행", type="primary", use_container_width=True)

# 세션 상태
if "analysis" not in st.session_state:
    st.session_state.analysis = None
if "analysis_id" not in st.session_state:
    st.session_state.analysis_id = None
if "excel_bytes" not in st.session_state:
    st.session_state.excel_bytes = None
if "excel_name" not in st.session_state:
    st.session_state.excel_name = None
if "excel_id" not in st.session_state:
    st.session_state.excel_id = None

current_id = None
if reading_file is not None:
    current_id = f"{getattr(reading_file, 'name', '')}:{getattr(reading_file, 'size', '')}"

# 업로드 파일이 바뀌면 결과 초기화
if st.session_state.analysis is not None and current_id and st.session_state.analysis_id != current_id:
    st.session_state.analysis = None
    st.session_state.analysis_id = None
    st.session_state.excel_bytes = None
    st.session_state.excel_name = None
    st.session_state.excel_id = None
    for k in ["cb_11", "cb_12", "cb_21", "cb_22", "cb_init_for_id"]:
        if k in st.session_state:
            del st.session_state[k]
    st.info("업로드 파일이 변경되었습니다. 실행 버튼을 다시 눌러 분석을 갱신해 주세요.")


def _analyze(uploaded) -> Dict[str, object]:
    filename = getattr(uploaded, "name", "")
    grade, cls, gc_text_raw = _extract_grade_class_from_filename_with_raw(filename)

    xls = pd.ExcelFile(uploaded, engine="openpyxl")
    sheet = xls.sheet_names[0]
    df_raw_top = xls.parse(sheet_name=sheet, header=None, dtype=str)

    if grade is None or cls is None:
        g2, c2 = _parse_grade_class_from_sheet_top(df_raw_top)
        grade = grade if grade is not None else g2
        cls = cls if cls is not None else c2
        if grade is not None and cls is not None:
            gc_text_raw = f"{grade}-{cls}"

    gc_text = gc_text_raw if gc_text_raw else "학년반미인식"

    req_2024_map = _load_required_books_from_repo(str(REQUIRED_2024_PATH))
    req_2025_map = _load_required_books_from_repo(str(REQUIRED_2025_PATH))

    df = _load_reading_table(xls, sheet)

    num_col = _pick_col(df, ["번호"])
    name_col = _pick_col(df, ["성명", "이름"])
    last_col = _pick_col(df, ["성"])
    first_col = _pick_col(df, ["명"])
    year_col = _pick_col(df, ["학년도"])
    sem_col = _pick_col(df, ["학기"])
    reading_col = _pick_col(df, ["독서활동상황", "독서활동 상황", "독서활동", "독서"])

    if not num_col:
        raise ValueError("독서활동 파일에서 '번호' 컬럼을 찾지 못했습니다.")
    if not sem_col:
        raise ValueError("독서활동 파일에서 '학기' 컬럼을 찾지 못했습니다.")
    if not reading_col:
        raise ValueError("독서활동 파일에서 '독서활동상황' 컬럼을 찾지 못했습니다.")

    if not year_col:
        inferred_year = _infer_academic_year_from_print_date(df_raw_top) or 2025
        df["__학년도__"] = str(inferred_year)
        year_col = "__학년도__"

    if not name_col:
        if last_col and first_col:
            df["__이름__"] = df[last_col].fillna("").astype(str).str.strip() + df[first_col].fillna("").astype(str).str.strip()
            name_col = "__이름__"
        else:
            raise ValueError("독서활동 파일에서 '성명/이름' 컬럼을 찾지 못했습니다.")

    df[num_col] = df[num_col].ffill()
    df[name_col] = df[name_col].ffill()
    df[year_col] = df[year_col].ffill()
    df[sem_col] = df[sem_col].ffill()

    df2 = df[[num_col, name_col, year_col, sem_col, reading_col]].copy()
    df2.columns = ["번호", "이름", "학년도", "학기", "독서활동상황"]
    df2["번호"] = df2["번호"].astype(str).str.extract(r"(\d+)")[0]
    df2 = df2.dropna(subset=["번호", "이름", "학년도", "학기"]).copy()

    groups = []
    for (num, name, year, sem), g in df2.groupby(["번호", "이름", "학년도", "학기"], dropna=False):
        all_books: List[Tuple[str, str]] = []
        for cell in g["독서활동상황"].tolist():
            all_books.extend(_split_books(cell))

        seen_in_group: List[dict] = []
        uniq_books: List[Tuple[str, str]] = []
        dup_in_group: List[Tuple[str, str, str]] = []

        for t, a in all_books:
            hit = _find_duplicate_against_seen(t, a, seen_in_group)
            if hit:
                it, ts, ascore = hit
                dup_in_group.append((t, a, f"유사중복(제목 {ts:.2f}, 저자 {ascore:.2f})"))
                continue
            seen_in_group.append({"title": t, "author": a})
            uniq_books.append((t, a))

        if grade is not None and cls is not None:
            sid = f"{int(grade)}{int(cls):02d}{int(num):02d}"
        else:
            sid = ""

        groups.append(
            {
                "번호": int(num),
                "이름": str(name).strip(),
                "학년도": str(year).strip(),
                "학기": str(sem).strip(),
                "학번": sid,
                "_books_raw": uniq_books,
                "_dup_in_group": dup_in_group,
            }
        )

    groups_sorted = sorted(groups, key=lambda x: (x["학번"], _semester_sort_key(x["학년도"], x["학기"]), x["번호"]))
    seen_by_student: Dict[str, List[dict]] = {}

    output_rows = []
    for item in groups_sorted:
        sid = item["학번"]
        student_key = sid or f"NOID-{item['번호']}-{item['이름']}"
        if student_key not in seen_by_student:
            seen_by_student[student_key] = []

        req_map = _choose_required_map(item["학년도"], req_2024_map, req_2025_map)
        req_keys = set(req_map.keys())
        req_title_map = _build_required_title_map(req_map)

        included: List[Tuple[str, str]] = []
        dup_remarks: List[str] = []
        review_remarks: List[str] = []

        for (t, a, reason) in item["_dup_in_group"]:
            dup_remarks.append(f"{t}({a}) 중복[{reason}]")

        for (t, a) in item["_books_raw"]:
            hit = _find_duplicate_against_seen(t, a, seen_by_student[student_key])
            if hit:
                it, ts, ascore = hit
                dup_remarks.append(
                    f"{t}({a}) 중복[학생전체 유사중복: 제목 {ts:.2f}, 저자 {ascore:.2f}] ↔ {it.get('title','')}({it.get('author','')})"
                )
                continue
            seen_by_student[student_key].append({"title": t, "author": a})
            included.append((t, a))

        books_for_cell: List[Tuple[str, str, bool]] = []
        required_count = 0

        for (t, a) in included:
            nt = _normalize_text(t)
            na = _normalize_text(a)
            exact_key = nt + "|" + na

            if exact_key in req_keys:
                required_count += 1
                books_for_cell.append((t, a, True))
                continue

            matched_std = None
            for k in _title_variants_norm(t):
                if k in req_title_map:
                    matched_std = req_title_map[k]
                    break

            if matched_std:
                std_t, std_a = matched_std
                a_score = _best_author_similarity(a, std_a)
                if a_score >= REQUIRED_AUTHOR_THRESHOLD:
                    required_count += 1
                    books_for_cell.append((t, a, True))
                    if _normalize_text(a) != _normalize_text(std_a):
                        review_remarks.append(
                            f"필독서 저자 표기 상이(확정, 저자 {a_score:.2f}): {t}({a}) → 필독: {std_t}({std_a})"
                        )
                    continue
                review_remarks.append(
                    f"필독서 제목 일치(검토, 저자 {a_score:.2f}): {t}({a}) → 후보: {std_t}({std_a})"
                )
                books_for_cell.append((t, a, False))
                continue

            fuzzy = _fuzzy_match_required_title_author(
                raw_title=t,
                raw_author=a,
                req_title_map=req_title_map,
                title_threshold=REQUIRED_TITLE_THRESHOLD,
                author_threshold=REQUIRED_AUTHOR_THRESHOLD,
            )
            if fuzzy:
                _, (std_t, std_a), tscore, ascore = fuzzy
                required_count += 1
                books_for_cell.append((t, a, True))
                review_remarks.append(
                    f"필독서 유사(확정, 제목 {tscore:.2f}/저자 {ascore:.2f}): {t}({a}) → 추정 필독: {std_t}({std_a})"
                )
                continue

            books_for_cell.append((t, a, False))

        total_count = len(included)
        satisfied = "충족" if (total_count >= 5 and required_count >= 1) else "미충족"

        remarks_all = sorted(set(dup_remarks + review_remarks))
        output_rows.append(
            {
                "학번": sid,
                "이름": item["이름"],
                "학년도": item["학년도"],
                "학기": item["학기"],
                "도서목록_표시": books_for_cell,
                "총권수": total_count,
                "필독서 권수": required_count,
                "충족 여부": satisfied,
                "비고": "\n".join(remarks_all),
            }
        )

    return {"gc_text": gc_text, "output_rows": output_rows}


if run:
    if not reading_file:
        st.error("독서활동상황 파일을 업로드해 주세요.")
        st.stop()

    prog = st.progress(0, text="분석 준비 중...")
    try:
        with st.spinner("분석을 진행하고 있습니다..."):
            prog.progress(10, text="엑셀 파일 읽는 중...")
            result = _analyze(reading_file)
            prog.progress(80, text="결과 정리 중...")

        st.session_state.analysis = result
        st.session_state.analysis_id = current_id
        st.session_state.excel_bytes = None
        st.session_state.excel_name = None
        st.session_state.excel_id = None
        for k in ["cb_11", "cb_12", "cb_21", "cb_22", "cb_init_for_id"]:
            if k in st.session_state:
                del st.session_state[k]

        prog.progress(100, text="완료")
        st.success("분석이 완료되었습니다. 아래에서 학기 기준을 조정해 주세요.")
    except Exception as e:
        prog.empty()
        st.error(f"분석 중 오류: {e}")
        st.stop()

analysis = st.session_state.analysis
if analysis is None:
    st.stop()

output_rows: List[dict] = analysis["output_rows"]  # type: ignore
gc_text: str = analysis["gc_text"]  # type: ignore

# =========================================================
# 5) 학기 선택(체크 해제 시 '-' 처리 + 상세/엑셀도 동일 반영)
# =========================================================
years_all = sorted({y for y in (_safe_year_int(r["학년도"]) for r in output_rows) if y is not None})
base_year = years_all[0] if years_all else None
next_year = years_all[1] if len(years_all) >= 2 else None

available_keys: set = set()
for r in output_rows:
    k = _row_semkey(r, base_year, next_year)
    if k:
        available_keys.add(k)

init_key = st.session_state.analysis_id
if st.session_state.get("cb_init_for_id") != init_key:
    st.session_state["cb_11"] = ("11" in available_keys)
    st.session_state["cb_12"] = ("12" in available_keys)
    st.session_state["cb_21"] = ("21" in available_keys)
    st.session_state["cb_22"] = ("22" in available_keys)
    st.session_state["cb_init_for_id"] = init_key

st.subheader("총 충족 기준(학기) 선택")
st.caption("체크 해제한 학기는 요약에서 '-'로 표시되며, 총 충족 여부 계산 및 상세/엑셀에서 제외됩니다.")

colA, colB = st.columns(2)
with colA:
    st.checkbox("1학년 1학기", key="cb_11", disabled=("11" not in available_keys))
    st.checkbox("1학년 2학기", key="cb_12", disabled=("12" not in available_keys))
with colB:
    st.checkbox("2학년 1학기", key="cb_21", disabled=("21" not in available_keys))
    st.checkbox("2학년 2학기", key="cb_22", disabled=("22" not in available_keys))

selected_keys: List[str] = []
if st.session_state.get("cb_11"):
    selected_keys.append("11")
if st.session_state.get("cb_12"):
    selected_keys.append("12")
if st.session_state.get("cb_21"):
    selected_keys.append("21")
if st.session_state.get("cb_22"):
    selected_keys.append("22")

# 선택이 바뀌면(다운로드 파일 불일치 방지) 생성된 엑셀 캐시 제거
excel_id = f"{st.session_state.analysis_id}:{','.join(selected_keys)}"
if st.session_state.excel_id is not None and st.session_state.excel_id != excel_id:
    st.session_state.excel_bytes = None
    st.session_state.excel_name = None
    st.session_state.excel_id = None
    st.info("학기 선택이 변경되었습니다. 엑셀 파일은 다시 생성해 주세요.")

if len(selected_keys) == 0:
    st.warning("선택된 학기가 없습니다. 총 충족 여부는 '판정 보류'로 표시되며, 상세/엑셀은 비어 있게 됩니다.")

# =========================================================
# 6) 요약 생성 (체크 해제 학기 = '-', 선택 학기만 O/X)
# =========================================================
def _default_value_for_sem(k: str) -> str:
    if k not in selected_keys:
        return "-"  # 선택 안 함
    return "X"  # 선택함(파일에 존재하는 학기만 선택 가능하도록 UI에서 제어)

summary_map: Dict[str, dict] = {}

for r in output_rows:
    sid = r["학번"]
    key = sid or f"NOID-{r['이름']}"
    if key not in summary_map:
        summary_map[key] = {
            "학번": sid,
            "이름": r["이름"],
            "11": _default_value_for_sem("11"),
            "12": _default_value_for_sem("12"),
            "21": _default_value_for_sem("21"),
            "22": _default_value_for_sem("22"),
        }

    semkey = _row_semkey(r, base_year, next_year)
    if not semkey or semkey not in selected_keys:
        continue  # 선택되지 않은 학기는 기록이 있어도 '-' 유지

    mark = "O" if r["충족 여부"] == "충족" else "X"
    summary_map[key][semkey] = mark

def _sort_key_summary(item: dict):
    sid = item.get("학번", "") or ""
    return (0, sid) if sid else (1, item.get("이름", ""))

summary_rows = sorted(summary_map.values(), key=_sort_key_summary)

summary_final = []
for i, row in enumerate(summary_rows, start=1):
    if len(selected_keys) == 0:
        total = "판정 보류"
    else:
        ok = True
        for k in selected_keys:
            if row.get(k, "-") != "O":
                ok = False
                break
        total = "충족" if ok else "미충족"

    summary_final.append(
        {
            "연번": i,
            "학번": row["학번"],
            "이름": row["이름"],
            "1학년 1학기 충족여부": row["11"],
            "1학년 2학기 충족여부": row["12"],
            "2학년 1학기 충족여부": row["21"],
            "2학년 2학기 충족여부": row["22"],
            "총 충족여부": total,
        }
    )

df_summary = pd.DataFrame(
    summary_final,
    columns=[
        "연번",
        "학번",
        "이름",
        "1학년 1학기 충족여부",
        "1학년 2학기 충족여부",
        "2학년 1학기 충족여부",
        "2학년 2학기 충족여부",
        "총 충족여부",
    ],
)

# =========================================================
# 7) 상세(선택 학기만 남김)
# =========================================================
detail_rows_selected = []
for r in output_rows:
    semkey = _row_semkey(r, base_year, next_year)
    if semkey and semkey in selected_keys:
        detail_rows_selected.append(r)

detail_rows_selected = sorted(detail_rows_selected, key=lambda x: (x.get("학번", ""), _semester_sort_key(x.get("학년도", ""), x.get("학기", "")), x.get("이름", "")))

st.subheader("요약 미리보기")
st.dataframe(df_summary.head(20), use_container_width=True)

st.subheader("상세 미리보기(상위 20행)")
df_detail_preview = pd.DataFrame(
    [
        {
            "연번": i,
            "학번": r["학번"],
            "이름": r["이름"],
            "학년도": r["학년도"],
            "학기": r["학기"],
            "총권수": r["총권수"],
            "필독서 권수": r["필독서 권수"],
            "충족 여부": r["충족 여부"],
            "비고": r["비고"],
        }
        for i, r in enumerate(detail_rows_selected, start=1)
    ]
)
st.dataframe(df_detail_preview.head(20), use_container_width=True)

# =========================================================
# 8) 엑셀 생성/다운로드: 생성 버튼 + 스피너/진행바 제공 (요청 2)
# =========================================================
def _build_excel_bytes(df_sum: pd.DataFrame, detail_rows: List[dict]) -> Tuple[bytes, str]:
    wb = Workbook()

    # 요약
    ws_sum = wb.active
    ws_sum.title = "요약"
    sum_headers = list(df_sum.columns)
    ws_sum.append(sum_headers)
    for _, row in df_sum.iterrows():
        ws_sum.append([row[h] for h in sum_headers])

    header_font = Font(bold=True)
    for c in range(1, len(sum_headers) + 1):
        cell = ws_sum.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_sum.column_dimensions["A"].width = 6
    ws_sum.column_dimensions["B"].width = 10
    ws_sum.column_dimensions["C"].width = 10
    ws_sum.column_dimensions["D"].width = 18
    ws_sum.column_dimensions["E"].width = 18
    ws_sum.column_dimensions["F"].width = 18
    ws_sum.column_dimensions["G"].width = 18
    ws_sum.column_dimensions["H"].width = 12

    # 상세(선택 학기만)
    ws = wb.create_sheet(title="상세")
    detail_headers = ["연번", "학번", "이름", "학년도", "학기", "도서명", "총권수", "필독서 권수", "충족 여부", "비고"]
    ws.append(detail_headers)

    for idx, r in enumerate(detail_rows, start=2):
        ws.append(
            [
                idx - 1,
                r["학번"],
                r["이름"],
                r["학년도"],
                r["학기"],
                "",
                r["총권수"],
                r["필독서 권수"],
                r["충족 여부"],
                r["비고"],
            ]
        )
        _set_books_cell(ws, row_idx=idx, col_idx=6, books=r["도서목록_표시"])

        note_cell = ws.cell(row=idx, column=10)
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")
        if "중복" in str(r.get("비고", "") or ""):
            note_cell.font = Font(color="FF0000")

    for c in range(1, len(detail_headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 70
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 55

    for rr in range(2, ws.max_row + 1):
        ws.cell(row=rr, column=6).alignment = Alignment(wrap_text=True, vertical="top")

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    sem_tag = "none" if len(selected_keys) == 0 else "-".join(selected_keys)
    filename = f"{gc_text}_독서가_결과_{sem_tag}.xlsx"
    return out.getvalue(), filename


st.divider()
st.subheader("엑셀 다운로드")

create_excel = st.button("📦 엑셀 파일 생성", use_container_width=True)
if create_excel:
    p = st.progress(0, text="엑셀 생성 준비 중...")
    with st.spinner("엑셀 파일을 생성하고 있습니다..."):
        p.progress(20, text="요약 시트 구성 중...")
        p.progress(50, text="상세 시트 구성 중...")
        # (선택 학기만 반영)
        bytes_data, fname = _build_excel_bytes(df_summary, detail_rows_selected)
        p.progress(100, text="완료")
    st.session_state.excel_bytes = bytes_data
    st.session_state.excel_name = fname
    st.session_state.excel_id = excel_id
    st.success("엑셀 파일 생성이 완료되었습니다. 아래 버튼으로 다운로드해 주세요.")

if st.session_state.excel_bytes is not None:
    st.download_button(
        label="📥 결과 엑셀 다운로드",
        data=st.session_state.excel_bytes,
        file_name=st.session_state.excel_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

if not RICH_TEXT_AVAILABLE:
    st.warning("현재 환경에서 '셀 내부 일부 굵게'가 제한되어, 필독서는 ★ 표시로 강조됩니다.")
